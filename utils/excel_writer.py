import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.logger import Logger

def format_cell_value(val) -> str:
    """
    Safely converts lists and dictionaries to clean, bulleted or line-separated strings
    so openpyxl does not crash when writing list/dict outputs from LLMs.
    """
    if val is None:
        return ""
    if isinstance(val, list):
        if not val:
            return ""
        if len(val) == 1:
            return str(val[0])
        return "\n".join(f"• {str(x).strip()}" for x in val)
    if isinstance(val, dict):
        return "\n".join(f"{k}: {v}" for k, v in val.items())
    return str(val)

def write_opportunities_to_excel(
    sponsored_trips: list, 
    ugc_deals: list, 
    speaker_hackathons: list, 
    output_dir: str = "output"
) -> str:
    """
    Writes opportunity data lists into a highly styled timestamped Excel file.
    Includes a Summary dashboard + 3 beautifully styled sheets.
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        Logger.info(f"Created output directory: '{output_dir}'")
        
    wb = Workbook()
    
    # ----------------------------------------------------
    # STYLING DEFINITIONS
    # ----------------------------------------------------
    # Color Palette: Deep Blue / Slate Theme
    COLOR_HEADER_BG = "1F385C"       # Deep Steel Blue
    COLOR_HEADER_FG = "FFFFFF"       # White
    COLOR_ALT_ROW_BG = "F4F7FA"      # Very Light Soft Slate Blue
    COLOR_TRAVEL_BG = "E2F0D9"       # Soft Sage Green for travel covered
    COLOR_TRAVEL_FG = "385723"       # Dark Olive Green for travel covered text
    COLOR_CARD_BG = "EBF1F5"         # Accent gray-blue for summary dashboard cards
    COLOR_BORDER = "D9D9D9"          # Soft light gray borders
    
    # Fonts
    font_title = Font(name="Calibri", size=18, bold=True, color="1F385C")
    font_section = Font(name="Calibri", size=13, bold=True, color="1F385C")
    font_header = Font(name="Calibri", size=11, bold=True, color=COLOR_HEADER_FG)
    font_data = Font(name="Calibri", size=10, bold=False, color="000000")
    font_data_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_italic = Font(name="Calibri", size=9, italic=True, color="595959")
    font_travel = Font(name="Calibri", size=10, bold=True, color=COLOR_TRAVEL_FG)
    
    # Fills
    fill_header = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    fill_alt_row = PatternFill(start_color=COLOR_ALT_ROW_BG, end_color=COLOR_ALT_ROW_BG, fill_type="solid")
    fill_travel = PatternFill(start_color=COLOR_TRAVEL_BG, end_color=COLOR_TRAVEL_BG, fill_type="solid")
    fill_card = PatternFill(start_color=COLOR_CARD_BG, end_color=COLOR_CARD_BG, fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Borders
    thin_border_side = Side(border_style="thin", color=COLOR_BORDER)
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom = Border(bottom=Side(border_style="medium", color="1F385C"))
    
    # Alignments
    align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_center_top = Alignment(horizontal="center", vertical="top", wrap_text=True)
    align_header = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_title = Alignment(horizontal="left", vertical="center")
    
    # ----------------------------------------------------
    # SHEET 0: SUMMARY DASHBOARD
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary Dashboard"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_summary.cell(row=2, column=2, value="OPPORTUNITY HUNTER DASHBOARD").font = font_title
    ws_summary.cell(row=2, column=2).alignment = align_title
    
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws_summary.cell(row=3, column=2, value=f"Generated on: {now_str} | Target Profile: Shriyash Soni (Developer, Speaker, Builder)").font = font_italic
    
    # Stat Summary Table
    ws_summary.cell(row=5, column=2, value="Metrics Summary").font = font_section
    ws_summary.row_dimensions[6].height = 25
    
    stats_headers = ["Opportunity Category", "Total Identified", "Fully Sponsored / Travel Covered"]
    for col_idx, text in enumerate(stats_headers, start=2):
        cell = ws_summary.cell(row=6, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_all
        
    categories_data = [
        ("Sponsored Trips (100% Free flight + stay)", len(sponsored_trips), sum(1 for x in sponsored_trips if format_cell_value(x.get("fully_covered")).lower() == "yes")),
        ("UGC Brand Deals (Paid creator collabs)", len(ugc_deals), "N/A"),
        ("Speaker Calls & Hackathons", len(speaker_hackathons), sum(1 for x in speaker_hackathons if format_cell_value(x.get("travel_covered")).lower() == "yes")),
    ]
    
    for row_idx, (cat_name, total, covered) in enumerate(categories_data, start=7):
        c1 = ws_summary.cell(row=row_idx, column=2, value=cat_name)
        c2 = ws_summary.cell(row=row_idx, column=3, value=total)
        c3 = ws_summary.cell(row=row_idx, column=4, value=covered)
        
        for c in (c1, c2, c3):
            c.font = font_data
            c.border = border_all
            c.alignment = align_top_left
            if row_idx % 2 == 1:
                c.fill = fill_alt_row
                
        c2.alignment = align_center_top
        c3.alignment = align_center_top
        
    # Top Picks Section
    ws_summary.cell(row=11, column=2, value="★ Curated High-Value Top Picks ★").font = font_section
    
    picks_headers = ["Category", "Opportunity / Event Name", "Host / Brand", "Compensation / Benefits Details", "Deadline", "Application Link"]
    ws_summary.row_dimensions[12].height = 25
    for col_idx, text in enumerate(picks_headers, start=2):
        cell = ws_summary.cell(row=12, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_all
        
    # Pick top 2 items from each category to highlight
    top_picks = []
    for item in sponsored_trips[:2]:
        top_picks.append((
            "Sponsored Trip", 
            format_cell_value(item.get("title")), 
            format_cell_value(item.get("organization")), 
            format_cell_value(item.get("benefits")), 
            format_cell_value(item.get("deadline")), 
            format_cell_value(item.get("url"))
        ))
    for item in ugc_deals[:2]:
        top_picks.append((
            "UGC Brand Deal", 
            format_cell_value(item.get("title")), 
            format_cell_value(item.get("brand")), 
            format_cell_value(item.get("compensation")), 
            format_cell_value(item.get("deadline")), 
            format_cell_value(item.get("url"))
        ))
    for item in speaker_hackathons[:2]:
        top_picks.append((
            "Speaker & Hackathon", 
            format_cell_value(item.get("event_name")), 
            format_cell_value(item.get("type")), 
            format_cell_value(item.get("compensation_or_prize")), 
            format_cell_value(item.get("deadline")), 
            format_cell_value(item.get("url"))
        ))
        
    curr_row = 13
    for pick in top_picks:
        ws_summary.row_dimensions[curr_row].height = 40  # generous height for wrapped lists
        c1 = ws_summary.cell(row=curr_row, column=2, value=pick[0])
        c2 = ws_summary.cell(row=curr_row, column=3, value=pick[1])
        c3 = ws_summary.cell(row=curr_row, column=4, value=pick[2])
        c4 = ws_summary.cell(row=curr_row, column=5, value=pick[3])
        c5 = ws_summary.cell(row=curr_row, column=6, value=pick[4])
        c6 = ws_summary.cell(row=curr_row, column=7)
        
        # Style link
        url = pick[5]
        if url and url.startswith("http"):
            c6.hyperlink = url
            c6.value = "Click to View Apply Link"
            c6.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        else:
            c6.value = url or "N/A"
            c6.font = font_italic
            
        for idx, cell in enumerate((c1, c2, c3, c4, c5, c6), start=2):
            if idx != 7: # Skip font reset for hyperlink
                cell.font = font_data_bold if idx == 2 else font_data
            cell.border = border_all
            cell.alignment = align_top_left
            
            # Soft sage green for travel-covered row highlights in Top Picks
            if pick[0] == "Sponsored Trip" or (pick[0] == "Speaker & Hackathon" and "reimbursed" in str(pick[3]).lower()):
                cell.fill = fill_travel
                if idx != 7:
                    cell.font = font_travel
            else:
                if curr_row % 2 == 1:
                    cell.fill = fill_alt_row
                    
        curr_row += 1
        
    # Auto-adjust column widths for Dashboard
    for col in ws_summary.iter_cols(min_col=2, max_col=7):
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.hyperlink:
                val_str = "Click to View Apply Link"
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws_summary.column_dimensions[col_letter].width = max(15, min(45, max_len + 3))
        
    # ----------------------------------------------------
    # SHEET 1: SPONSORED TRIPS
    # ----------------------------------------------------
    ws_trips = wb.create_sheet(title="Sponsored Trips")
    ws_trips.views.sheetView[0].showGridLines = True
    
    trips_headers = [
        "Title of Opportunity", "Hosting Organization", "Benefits & Coverage Details", 
        "Fully Paid Flight+Hotel?", "Eligibility & Requirements", "Deadline", "Application URL", "Source Query"
    ]
    
    ws_trips.row_dimensions[1].height = 28
    for col_idx, text in enumerate(trips_headers, start=1):
        cell = ws_trips.cell(row=1, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_all
        
    for r_idx, opp in enumerate(sponsored_trips, start=2):
        ws_trips.row_dimensions[r_idx].height = 45
        
        is_fully = format_cell_value(opp.get("fully_covered")).strip().lower() == "yes"
        row_fill = fill_travel if is_fully else (fill_alt_row if r_idx % 2 == 1 else fill_white)
        row_font = font_travel if is_fully else font_data
        
        c1 = ws_trips.cell(row=r_idx, column=1, value=format_cell_value(opp.get("title")))
        c2 = ws_trips.cell(row=r_idx, column=2, value=format_cell_value(opp.get("organization")))
        c3 = ws_trips.cell(row=r_idx, column=3, value=format_cell_value(opp.get("benefits")))
        c4 = ws_trips.cell(row=r_idx, column=4, value=format_cell_value(opp.get("fully_covered")))
        c5 = ws_trips.cell(row=r_idx, column=5, value=format_cell_value(opp.get("eligibility")))
        c6 = ws_trips.cell(row=r_idx, column=6, value=format_cell_value(opp.get("deadline")))
        c7 = ws_trips.cell(row=r_idx, column=7)
        c8 = ws_trips.cell(row=r_idx, column=8, value=format_cell_value(opp.get("source_query")))
        
        # Link styling
        url = format_cell_value(opp.get("url"))
        if url and url.startswith("http"):
            c7.hyperlink = url
            c7.value = url
            c7.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        else:
            c7.value = url or "N/A"
            c7.font = font_italic
            
        c4.alignment = align_center_top
        c6.alignment = align_center_top
        
        for c in (c1, c2, c3, c4, c5, c6, c7, c8):
            c.border = border_all
            if c != c7:
                c.font = row_font
            c.fill = row_fill
            if c != c4 and c != c6:
                c.alignment = align_top_left
                
    # ----------------------------------------------------
    # SHEET 2: UGC BRAND DEALS
    # ----------------------------------------------------
    ws_ugc = wb.create_sheet(title="UGC Brand Deals")
    ws_ugc.views.sheetView[0].showGridLines = True
    
    ugc_headers = [
        "Brand/Company", "Campaign/Deal Title", "Product & Campaign Details", 
        "Confirmed Compensation", "Creator Requirements", "Deadline", "Application URL", "Source Query"
    ]
    
    ws_ugc.row_dimensions[1].height = 28
    for col_idx, text in enumerate(ugc_headers, start=1):
        cell = ws_ugc.cell(row=1, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_all
        
    for r_idx, opp in enumerate(ugc_deals, start=2):
        ws_ugc.row_dimensions[r_idx].height = 45
        row_fill = fill_alt_row if r_idx % 2 == 1 else fill_white
        
        c1 = ws_ugc.cell(row=r_idx, column=1, value=format_cell_value(opp.get("brand")))
        c2 = ws_ugc.cell(row=r_idx, column=2, value=format_cell_value(opp.get("title")))
        c3 = ws_ugc.cell(row=r_idx, column=3, value=format_cell_value(opp.get("description")))
        c4 = ws_ugc.cell(row=r_idx, column=4, value=format_cell_value(opp.get("compensation")))
        c5 = ws_ugc.cell(row=r_idx, column=5, value=format_cell_value(opp.get("requirements")))
        c6 = ws_ugc.cell(row=r_idx, column=6, value=format_cell_value(opp.get("deadline")))
        c7 = ws_ugc.cell(row=r_idx, column=7)
        c8 = ws_ugc.cell(row=r_idx, column=8, value=format_cell_value(opp.get("source_query")))
        
        # Link styling
        url = format_cell_value(opp.get("url"))
        if url and url.startswith("http"):
            c7.hyperlink = url
            c7.value = url
            c7.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        else:
            c7.value = url or "N/A"
            c7.font = font_italic
            
        c6.alignment = align_center_top
        
        for c in (c1, c2, c3, c4, c5, c6, c7, c8):
            c.border = border_all
            if c != c7:
                c.font = font_data
            c.fill = row_fill
            if c != c6:
                c.alignment = align_top_left
                
    # ----------------------------------------------------
    # SHEET 3: SPEAKER & HACKATHONS
    # ----------------------------------------------------
    ws_hack = wb.create_sheet(title="Speaker Calls & Hackathons")
    ws_hack.views.sheetView[0].showGridLines = True
    
    hack_headers = [
        "Event/Conference Name", "Opportunity Type", "Event Themes/Description", 
        "Prizes & Stipends", "Travel Reimbursed?", "Location & Dates", "Deadline", "Apply URL", "Source Query"
    ]
    
    ws_hack.row_dimensions[1].height = 28
    for col_idx, text in enumerate(hack_headers, start=1):
        cell = ws_hack.cell(row=1, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_all
        
    for r_idx, opp in enumerate(speaker_hackathons, start=2):
        ws_hack.row_dimensions[r_idx].height = 45
        
        has_travel = format_cell_value(opp.get("travel_covered")).strip().lower() == "yes"
        row_fill = fill_travel if has_travel else (fill_alt_row if r_idx % 2 == 1 else fill_white)
        row_font = font_travel if has_travel else font_data
        
        c1 = ws_hack.cell(row=r_idx, column=1, value=format_cell_value(opp.get("event_name")))
        c2 = ws_hack.cell(row=r_idx, column=2, value=format_cell_value(opp.get("type")))
        c3 = ws_hack.cell(row=r_idx, column=3, value=format_cell_value(opp.get("description")))
        c4 = ws_hack.cell(row=r_idx, column=4, value=format_cell_value(opp.get("compensation_or_prize")))
        c5 = ws_hack.cell(row=r_idx, column=5, value=format_cell_value(opp.get("travel_covered")))
        c6 = ws_hack.cell(row=r_idx, column=6, value=format_cell_value(opp.get("location_and_dates")))
        c7 = ws_hack.cell(row=r_idx, column=7, value=format_cell_value(opp.get("deadline")))
        c8 = ws_hack.cell(row=r_idx, column=8)
        c9 = ws_hack.cell(row=r_idx, column=9, value=format_cell_value(opp.get("source_query")))
        
        # Link styling
        url = format_cell_value(opp.get("url"))
        if url and url.startswith("http"):
            c8.hyperlink = url
            c8.value = url
            c8.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        else:
            c8.value = url or "N/A"
            c8.font = font_italic
            
        c2.alignment = align_center_top
        c5.alignment = align_center_top
        c7.alignment = align_center_top
        
        for c in (c1, c2, c3, c4, c5, c6, c7, c8, c9):
            c.border = border_all
            if c != c8:
                c.font = row_font
            c.fill = row_fill
            if c not in (c2, c5, c7):
                c.alignment = align_top_left
                
    # ----------------------------------------------------
    # DYNAMIC LAYOUT & COLUMN WIDTH ADJUSTMENTS FOR DATA SHEETS
    # ----------------------------------------------------
    # Set Auto-Filter, Freeze Panes, and Auto-Width for all three data sheets
    data_sheets = [ws_trips, ws_ugc, ws_hack]
    for ws in data_sheets:
        # Freeze first row (headers)
        ws.freeze_panes = "A2"
        
        # Enable Auto Filter
        last_col_letter = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"
        
        # Adjust widths dynamically
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                # Don't let hyperlinked long URLs bloat the column width, cap URL width calculation at 25
                if cell.hyperlink:
                    val_str = "http://application-link-url.com"
                if len(val_str) > max_len:
                    max_len = len(val_str)
            # Cap at 50, minimum 13
            ws.column_dimensions[col_letter].width = max(13, min(50, max_len + 3))
            
    # Save timestamped workbook
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"opportunities_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    wb.save(filepath)
    Logger.success(f"Master Workbook successfully saved to {os.path.abspath(filepath)}")
    
    # ----------------------------------------------------
    # ALSO SAVE SEPARATE INDIVIDUAL EXCEL FILES FOR EACH CATEGORY
    # ----------------------------------------------------
    try:
        from openpyxl import load_workbook
        
        # 1. Save Sponsored Trips only
        wb_trips = load_workbook(filepath)
        for sheet_name in ["Summary Dashboard", "UGC Brand Deals", "Speaker Calls & Hackathons"]:
            if sheet_name in wb_trips.sheetnames:
                wb_trips.remove(wb_trips[sheet_name])
        wb_trips.save(os.path.join(output_dir, f"sponsored_trips_{timestamp}.xlsx"))
        Logger.success(f"Individual Sponsored Trips workbook saved to output/sponsored_trips_{timestamp}.xlsx")
        
        # 2. Save UGC Brand Deals only
        wb_ugc = load_workbook(filepath)
        for sheet_name in ["Summary Dashboard", "Sponsored Trips", "Speaker Calls & Hackathons"]:
            if sheet_name in wb_ugc.sheetnames:
                wb_ugc.remove(wb_ugc[sheet_name])
        wb_ugc.save(os.path.join(output_dir, f"ugc_brand_deals_{timestamp}.xlsx"))
        Logger.success(f"Individual UGC Brand Deals workbook saved to output/ugc_brand_deals_{timestamp}.xlsx")
        
        # 3. Save Speaker & Hackathons only
        wb_hack = load_workbook(filepath)
        for sheet_name in ["Summary Dashboard", "Sponsored Trips", "UGC Brand Deals"]:
            if sheet_name in wb_hack.sheetnames:
                wb_hack.remove(wb_hack[sheet_name])
        wb_hack.save(os.path.join(output_dir, f"speaker_hackathons_{timestamp}.xlsx"))
        Logger.success(f"Individual Speaker & Hackathons workbook saved to output/speaker_hackathons_{timestamp}.xlsx")
        
    except Exception as exc:
        Logger.error(f"Error generating individual category Excel files: {str(exc)}")
        
    return filepath
